import os
import openpyxl
from openpyxl.styles import PatternFill
import shutil

def scan_for_initial_assesment(batch):

    #create the path from the batch given
    path='T:\data\R/' + batch

    #scan the directory and locate the .xlsx file
    for filename in os.listdir(path):
        if '.xlsx' and 'template' in filename: #make sure you always use a template excel file that NEVER changes name
            optical_yield_template=filename

    #join together to get a full directory string
    full_directory=os.path.join(path,optical_yield_template)

    #Open the Excel file
    workbook=openpyxl.load_workbook(full_directory)
    #reading the data from the first sheet
    sheet=workbook.active

    completed_panels = []
    non_completed_panels = []
    # Access and scan through the data in the Excel file
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[4] is None:
            #make sure the values in the list are unique
            if row[2] not in non_completed_panels:
                non_completed_panels.append(row[2])
        else:
            #make sure the values in the list are unique
            if row[2] not in completed_panels:
                completed_panels.append(row[2])

    return completed_panels, non_completed_panels

def scan_images_directory(batch):

    completed_panels = []

    path = 'T:\data\R/' + batch + '/optical/images'

    for root, dirs, files in os.walk(path):
        for name in files:

            #look for the panel number without the 0 contained in the filename (example: panel 3 not panel 03)
            #SOP make sure we don't get confused using panel 0X instead of panel X
            if name[6]=='1' or name[6]=='2' or name[6]=='3' or name[6]=='4' or name[6]=='5' or name[6]=='6':
                if name[6] not in completed_panels:
                    completed_panels.append(name[6])
            else:
                continue

    return completed_panels

def scan_resistance_check(batch):

    #create the path from the batch given
    path='T:\data\R/' + batch + '/electrical'

    #scan the directory and locate the .xlsx file
    for filename in os.listdir(path):
        if '.xlsx' and 'Resistance_checks' in filename: #make sure you always use a template excel file that NEVER changes name
            resistance_checks=filename
            print(resistance_checks)

    #join together to get a full directory string
    full_directory=os.path.join(path,resistance_checks)

    #Open the Excel file
    workbook=openpyxl.load_workbook(full_directory)
    #reading the data from the first sheet
    sheet=workbook.active


    panels_done=[]
    cells_done=[]
    # Access and scan through the data in the Excel file
    for row in sheet.iter_rows(min_row=2, values_only=True):
        #consider if this need to ber or instead of and
        if row[3] and row[4] and row[5] is not None:
            panel_ID=row[1]
            cell_ID=row[2]

            print('res_panel_ID',panel_ID)
            print('res_cell_ID',cell_ID)

            panels_done.append(panel_ID)
            cells_done.append(cell_ID)

    # Creating the dictionary with variable-sized tuples
    dictionary = {}
    current_key = None
    current_values = []

    for panel, cell in zip(panels_done, cells_done):
        if panel != current_key:
            # If a new panel is encountered, update the current key and values
            current_key = panel
            current_values = [cell]
        else:
            # If the same panel is encountered, append the cell to the values
            current_values.append(cell)

        # Update the dictionary with the current key and values
        dictionary[current_key] = tuple(current_values)

    return dictionary

def scan_for_direct_drive_images(batch):

    completed_panels = []

    path = 'T:\data\R/' + batch + '/optical/DIRECT DRIVE IMAGES'
    panels_done=[]
    cells_done=[]

    for root, dirs, files in os.walk(path):
        for name in files:
            panel_ID=name[5:7]
            cell_ID=name[8:10]

            panels_done.append(panel_ID)
            cells_done.append(cell_ID)

    # Creating the dictionary with variable-sized tuples
    dictionary = {}
    current_key = None
    current_values = []

    for panel, cell in zip(panels_done, cells_done):
        if panel != current_key:
            # If a new panel is encountered, update the current key and values
            current_key = panel
            current_values = [cell]
        else:
            # If the same panel is encountered, append the cell to the values
            current_values.append(cell)

        # Update the dictionary with the current key and values
        dictionary[current_key] = tuple(current_values)
    # Assuming you have already created the 'dictionary' using the previous code

    # Create a new dictionary to store unique values
    unique_dictionary = {}

    # Iterate through the original dictionary
    for key, values in dictionary.items():
        unique_values = []
        seen_values = set()

        # Iterate through the values in reverse order to keep the last occurrence
        for value in reversed(values):
            if value not in seen_values:
                unique_values.insert(0, value)  # Insert at the beginning to maintain order
                seen_values.add(value)

        unique_dictionary[key] = tuple(unique_values)

    return unique_dictionary

def scan_for_tp_connection_test(batch):

    #create the path from the batch given
    path='T:\data\R/' + batch + '/electrical'

    #scan the directory and locate the .xlsx file
    for filename in os.listdir(path):
        if '.xlsx' and 'TP_connection_test' in filename: #make sure you always use a template excel file that NEVER changes name
            resistance_checks=filename

    #join together to get a full directory string
    full_directory=os.path.join(path,resistance_checks)

    #Open the Excel file
    workbook=openpyxl.load_workbook(full_directory)
    #reading the data from the first sheet
    sheet=workbook.active

    panels_done=[]
    cells_done=[]
    # Access and scan through the data in the Excel file
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[3] is not None:
            panels_done.append(row[1])
            cells_done.append(row[2])

    # Creating the dictionary with variable-sized tuples
    dictionary = {}
    current_key = None
    current_values = []

    for panel, cell in zip(panels_done, cells_done):
        if panel != current_key:
            # If a new panel is encountered, update the current key and values
            current_key = panel
            current_values = [cell]
        else:
            # If the same panel is encountered, append the cell to the values
            current_values.append(cell)

        # Update the dictionary with the current key and values
        dictionary[current_key] = tuple(current_values)
    # Assuming you have already created the 'dictionary' using the previous code

    # Create a new dictionary to store unique values
    unique_dictionary = {}

    # Iterate through the original dictionary
    for key, values in dictionary.items():
        unique_values = []
        seen_values = set()

        # Iterate through the values in reverse order to keep the last occurrence
        for value in reversed(values):
            if value not in seen_values:
                unique_values.insert(0, value)  # Insert at the beginning to maintain order
                seen_values.add(value)

        unique_dictionary[key] = tuple(unique_values)

    return unique_dictionary

def copy_excel_files(source_dir, destination_dir):
    # Check if both source and destination directories exist
    if not os.path.exists(source_dir):
        print(f"Source directory '{source_dir}' does not exist.")
        return
    if not os.path.exists(destination_dir):
        print(f"Destination directory '{destination_dir}' does not exist.")
        return

    # Get a list of files in the source directory
    files = os.listdir(source_dir)

    # Copy Excel files one by one
    for file in files:
        if file.endswith('.xlsx') or file.endswith('.xls'):
            source_file = os.path.join(source_dir, file)
            destination_file = os.path.join(destination_dir, file)
            if not os.path.exists(destination_file):
                shutil.copyfile(source_file, destination_file)
                print(f"Copied '{source_file}' to '{destination_file}'")
            else:
                print(f"File '{destination_file}' already exists in the destination directory. Skipping.")

#Press the green button in the gutter to run the script.
if __name__ == '__main__':

    batch='R539'

    #copy excel files in directory if they are not there
    #resistance + TP_Connection
    source_directory = r"C:\Users\Harry.Delalis\PycharmProjects\moonfish_tracker\template_tests/electrical"

    destination_directory = r"T:\data\R/" +batch +"/electrical"
    copy_excel_files(source_directory, destination_directory)

    #copy excel files in directory if they are not there
    #initial_assessment
    source_directory = r"C:\Users\Harry.Delalis\PycharmProjects\moonfish_tracker\template_tests\initial_assessment"

    destination_directory = r"T:\data\R/" +batch
    copy_excel_files(source_directory, destination_directory)

    #copy the batch tracker itself in the data/R folder
    source_directory= r'C:\Users\Harry.Delalis\PycharmProjects\moonfish_tracker'
    destination_directory = r"T:\data\R/" + batch
    copy_excel_files(source_directory, destination_directory)



    #first_step
    completed_initial_assesssment,not_completed_initial_assesssment=scan_for_initial_assesment(batch)
    #second_step
    completed_global_motherglass_images=scan_images_directory(batch)
    #third_step
    completed_resistance_checks=scan_resistance_check(batch)
    #fourth_step
    completed_direct_drive_images=scan_for_direct_drive_images(batch)
    #fifth_step
    completed_tp_scan=scan_for_tp_connection_test(batch)


    print()
    print('First step: Initial Assessment')
    print('Panels that have been through initial assessment',completed_initial_assesssment)
    print('Panels that have not been through initial assessment',not_completed_initial_assesssment)
    print()
    print('Second step: Global mother glass images')
    print('Panels that have global images taken',completed_global_motherglass_images)
    print()
    print('Third step: Resistance checks')
    print('Panels and cells that have resistance taken',completed_resistance_checks)
    print()
    print('Fourth step: Direct drive images')
    print('Panels and cells that have direct drive images taken',completed_direct_drive_images)
    print()
    print('Fifth step: TP connection test')
    print('Panels and cells that have TP connection tests taken',completed_tp_scan)

    #Populate the excel file

    input_file=r'T:\data\R\R539\Moonfish Batch Lifetime.xlsx'
    output_file=r'T:\data\R\R539\Moonfish Batch Lifetime.xlsx'
    workbook = openpyxl.load_workbook(input_file)

    # Select the desired sheet (assuming it's the first sheet, change if needed)
    sheet = workbook.active

    for row in range(15,21):
        cell_value=sheet[f'D{row}'].value

        if cell_value in completed_initial_assesssment:
            sheet[f'D{row}'].fill=PatternFill(start_color='FF00FF00', end_color='FF00FF00', fill_type="solid")

    integer_list_images=[int(x) for x in completed_global_motherglass_images]

    for row in range(15,21):
        cell_value=sheet[f'F{row}'].value

        if cell_value in integer_list_images:
            sheet[f'F{row}'].fill=PatternFill(start_color='FF00FF00', end_color='FF00FF00', fill_type="solid")

    #Resistance checks and populate the excel

    excel_coordinates=((15,20),(20,25),(25,30),(30,35),(35,40),(40,45))
    count=1
    for j in excel_coordinates:
        try:
            for row in range(*j):
                cell_value_k_row=sheet[f'K{row}'].value
                cell_value_l_row=sheet[f'L{row}'].value
                cell_value_m_row=sheet[f'M{row}'].value
                cell_value_n_row=sheet[f'N{row}'].value

                if cell_value_n_row in completed_resistance_checks[count]:
                    sheet[f'N{row}'].fill = PatternFill(start_color='FF00FF00', end_color='FF00FF00', fill_type="solid")
                if cell_value_m_row in completed_resistance_checks[count]:
                    sheet[f'M{row}'].fill = PatternFill(start_color='FF00FF00', end_color='FF00FF00', fill_type="solid")
                if cell_value_l_row in completed_resistance_checks[count]:
                    sheet[f'L{row}'].fill = PatternFill(start_color='FF00FF00', end_color='FF00FF00', fill_type="solid")
                if cell_value_k_row in completed_resistance_checks[count]:
                    sheet[f'K{row}'].fill = PatternFill(start_color='FF00FF00', end_color='FF00FF00', fill_type="solid")
            count+=1
        except KeyError:
            print('panel not done continue')
            count+=1


    #Direct Drive images populate

    #convert to integer equivalent

    excel_coordinates=((15,20),(20,25),(25,30),(30,35),(35,40),(40,45))
    count=1
    iteration_list=('01','02','03','04','05','06')
    for j in excel_coordinates:
        try:
            values_for_first_key=completed_direct_drive_images[iteration_list[count-1]]

            integer_equivalents=[int(value) for value in values_for_first_key]

            for row in range(*j):
                cell_value_p_row=sheet[f'P{row}'].value
                cell_value_s_row=sheet[f'S{row}'].value

                if cell_value_p_row in integer_equivalents:
                    sheet[f'P{row}'].fill = PatternFill(start_color='FF00FF00', end_color='FF00FF00', fill_type="solid")
                if cell_value_s_row in integer_equivalents:
                    sheet[f'S{row}'].fill = PatternFill(start_color='FF00FF00', end_color='FF00FF00', fill_type="solid")

            count+=1
        except KeyError:
            print('panel missing continue')
            count+=1

    #TP_connection populate excel

    excel_coordinates=((15,20),(20,25),(25,30),(30,35),(35,40),(40,45))
    count=1
    for j in excel_coordinates:
        try:
            for row in range(*j):
                cell_value_y_row=sheet[f'Y{row}'].value
                cell_value_z_row=sheet[f'Z{row}'].value
                cell_value_aa_row=sheet[f'AA{row}'].value
                cell_value_ab_row=sheet[f'AB{row}'].value

                if cell_value_y_row in completed_tp_scan[count]:
                    sheet[f'Y{row}'].fill = PatternFill(start_color='FF00FF00', end_color='FF00FF00', fill_type="solid")
                if cell_value_z_row in completed_tp_scan[count]:
                    sheet[f'Z{row}'].fill = PatternFill(start_color='FF00FF00', end_color='FF00FF00', fill_type="solid")
                if cell_value_aa_row in completed_tp_scan[count]:
                    sheet[f'AA{row}'].fill = PatternFill(start_color='FF00FF00', end_color='FF00FF00', fill_type="solid")
                if cell_value_ab_row in completed_tp_scan[count]:
                    sheet[f'AB{row}'].fill = PatternFill(start_color='FF00FF00', end_color='FF00FF00', fill_type="solid")
            count+=1
        except KeyError:
            print('panel not done continue')
            count+=1

    # Save the modified workbook to a new file
    workbook.save(output_file)




